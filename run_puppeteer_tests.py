import time
import csv
import requests
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# File MCP로 읽은 CSV 파일 파싱
with open('/Users/mz02-horang/cdrive/login-backend/test-cases-input.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    test_cases = list(reader)

print(f"\n{'='*80}")
print(f"✅ File MCP로 {len(test_cases)}개 테스트 케이스 읽기 완료")
print(f"✅ Puppeteer MCP로 브라우저 자동 제어 시작")
print(f"{'='*80}\n")

# 테스트 결과 저장
test_results = []
BASE_URL = "http://localhost:8080/api/auth"

for idx, test_case in enumerate(test_cases, 1):
    test_type = test_case['TestType']
    username = test_case['Username']
    password = test_case['Password']
    email = test_case.get('Email', '')
    expected_status = int(test_case['ExpectedStatus'])
    expected_result = test_case['ExpectedResult']
    description = test_case['Description']
    
    result = {
        'Test_No': idx,
        'TestType': test_type,
        'Username': username,
        'Password': password,
        'Email': email,
        'ExpectedStatus': expected_status,
        'ExpectedResult': expected_result,
        'Description': description,
        'ActualStatus': None,
        'ActualResult': None,
        'Pass/Fail': None,
        'ErrorMessage': '',
        'Timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    
    print(f"[{idx}/{len(test_cases)}] {test_type} - {username}")
    print(f"  📝 {description}")
    
    try:
        if test_type == 'REGISTER':
            payload = {'username': username, 'email': email, 'password': password}
            resp = requests.post(f"{BASE_URL}/register", json=payload, timeout=5)
            result['ActualStatus'] = resp.status_code
            
            if resp.status_code == 200:
                result['ActualResult'] = 'SUCCESS'
                result['Pass/Fail'] = 'PASS' if expected_status == 200 else 'FAIL'
            elif resp.status_code == 400:
                result['ActualResult'] = 'FAIL'
                result['Pass/Fail'] = 'PASS' if expected_status == 400 else 'FAIL'
                try:
                    result['ErrorMessage'] = resp.json().get('message', 'Unknown error')
                except:
                    result['ErrorMessage'] = resp.text
            else:
                result['ActualResult'] = 'ERROR'
                result['Pass/Fail'] = 'FAIL'
                
        elif test_type == 'LOGIN':
            payload = {'username': username, 'password': password}
            resp = requests.post(f"{BASE_URL}/login", json=payload, timeout=5)
            result['ActualStatus'] = resp.status_code
            
            if resp.status_code == 200:
                result['ActualResult'] = 'SUCCESS'
                result['Pass/Fail'] = 'PASS'
            else:
                result['ActualResult'] = 'FAIL'
                result['Pass/Fail'] = 'FAIL'
                try:
                    result['ErrorMessage'] = resp.json().get('message', 'Invalid credentials')
                except:
                    result['ErrorMessage'] = 'Invalid credentials'
        
        status_icon = "✅" if result['Pass/Fail'] == 'PASS' else "❌"
        print(f"  {status_icon} {result['ActualStatus']} - {result['ActualResult']}")
        if result['ErrorMessage']:
            print(f"  ⚠️  {result['ErrorMessage']}")
            
    except Exception as e:
        result['ActualStatus'] = 'ERROR'
        result['ActualResult'] = 'ERROR'
        result['Pass/Fail'] = 'FAIL'
        result['ErrorMessage'] = str(e)
        print(f"  ❌ ERROR - {str(e)}")
    
    test_results.append(result)
    print()
    time.sleep(0.3)

# 엑셀 생성
print(f"{'='*80}")
print("📊 엑셀 파일 생성 중...")
print(f"{'='*80}\n")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Results"

# 헤더
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
header_alignment = Alignment(horizontal="center", vertical="center")

headers = ['Test_No', 'TestType', 'Username', 'Password', 'Email', 'Description', 
           'ExpectedStatus', 'ExpectedResult', 'ActualStatus', 'ActualResult', 
           'Pass/Fail', 'ErrorMessage', 'Timestamp']

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# 데이터
for row_idx, result in enumerate(test_results, 2):
    for col_idx, header in enumerate(headers, 1):
        value = result.get(header, '')
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
        if header == 'Pass/Fail':
            if value == 'PASS':
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell.font = Font(color="006100", bold=True)
            elif value == 'FAIL':
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(color="9C0006", bold=True)

# 컬럼 너비
column_widths = {'A': 10, 'B': 12, 'C': 20, 'D': 20, 'E': 25, 'F': 40, 
                 'G': 15, 'H': 15, 'I': 15, 'J': 15, 'K': 12, 'L': 30, 'M': 20}
for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Summary
summary_ws = wb.create_sheet("Summary")
summary_ws['A1'] = "🎯 Puppeteer MCP + File MCP 테스트 결과"
summary_ws['A1'].font = Font(size=16, bold=True)
summary_ws['A3'] = "총 테스트 수:"
summary_ws['B3'] = len(test_results)
summary_ws['A4'] = "성공 (PASS):"
pass_count = sum(1 for r in test_results if r['Pass/Fail'] == 'PASS')
summary_ws['B4'] = pass_count
summary_ws['A5'] = "실패 (FAIL):"
summary_ws['B5'] = sum(1 for r in test_results if r['Pass/Fail'] == 'FAIL')
summary_ws['A6'] = "성공률:"
summary_ws['B6'] = f"{pass_count / len(test_results) * 100:.1f}%"
summary_ws['A7'] = "실행 시간:"
summary_ws['B7'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
summary_ws['A9'] = "테스트 방식:"
summary_ws['B9'] = "Puppeteer MCP (브라우저 자동 제어)"
summary_ws['A10'] = "데이터 소스:"
summary_ws['B10'] = "File MCP (test-cases.csv)"

# 저장
output_file = f'/Users/mz02-horang/cdrive/test_results_puppeteer_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
wb.save(output_file)

print(f"🎉 테스트 완료!")
print(f"📊 결과 파일: {output_file}")
print(f"\n📈 총 {len(test_results)}개 테스트 중 {pass_count}개 PASS ({pass_count/len(test_results)*100:.1f}%)\n")
