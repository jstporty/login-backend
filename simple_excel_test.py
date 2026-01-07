import csv
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# File MCP로 읽은 CSV 파싱
with open('/Users/mz02-horang/cdrive/login-backend/test-cases-input.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    test_cases = list(reader)

print(f"\n✅ File MCP로 {len(test_cases)}개 테스트 케이스 읽기 완료\n")

# 테스트 결과
results = []
BASE_URL = "http://localhost:8080/api/auth"

for idx, tc in enumerate(test_cases, 1):
    test_type = tc['TestType']
    username = tc['Username']
    password = tc['Password']
    email = tc.get('Email', '')
    expected = tc['ExpectedResult']
    desc = tc['Description']
    
    print(f"[{idx}/10] {test_type} - {username}")
    
    try:
        if test_type == 'REGISTER':
            resp = requests.post(f"{BASE_URL}/register", 
                               json={'username': username, 'email': email, 'password': password})
        else:  # LOGIN
            resp = requests.post(f"{BASE_URL}/login",
                               json={'username': username, 'password': password})
        
        actual = "성공" if resp.status_code == 200 else "실패"
        result = "✅ PASS" if (expected == "SUCCESS" and actual == "성공") or (expected == "FAIL" and actual == "실패") else "❌ FAIL"
        
        results.append({
            'No': idx,
            'TestType': '회원가입' if test_type == 'REGISTER' else '로그인',
            'Username': username,
            'Description': desc,
            'Expected': '성공' if expected == 'SUCCESS' else '실패',
            'Actual': actual,
            'Result': result
        })
        
        print(f"  {result}: 예상={expected}, 실제={actual}\n")
        
    except Exception as e:
        results.append({
            'No': idx,
            'TestType': '회원가입' if test_type == 'REGISTER' else '로그인',
            'Username': username,
            'Description': desc,
            'Expected': '성공' if expected == 'SUCCESS' else '실패',
            'Actual': f"에러: {str(e)}",
            'Result': "❌ FAIL"
        })
        print(f"  ❌ FAIL: {e}\n")

# 간단한 엑셀 생성
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "테스트 결과"

# 헤더
headers = ['No', 'TestType', 'Username', 'Description', 'Expected', 'Actual', 'Result']
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 데이터
for row_idx, result in enumerate(results, 2):
    for col_idx, header in enumerate(headers, 1):
        value = result[header]
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="center" if col_idx <= 3 else "left", vertical="center")
        
        # Result 컬럼 색상
        if header == 'Result':
            if 'PASS' in value:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell.font = Font(color="006100", bold=True, size=11)
            else:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(color="9C0006", bold=True, size=11)

# 컬럼 너비
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 45
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 15

# 요약 추가
summary_row = len(results) + 3
ws.merge_cells(f'A{summary_row}:C{summary_row}')
ws[f'A{summary_row}'] = "📊 테스트 결과 요약"
ws[f'A{summary_row}'].font = Font(bold=True, size=14)
ws[f'A{summary_row}'].alignment = Alignment(horizontal="center")

pass_count = sum(1 for r in results if 'PASS' in r['Result'])
ws[f'A{summary_row+1}'] = f"총 테스트: {len(results)}개"
ws[f'A{summary_row+2}'] = f"성공: {pass_count}개"
ws[f'A{summary_row+3}'] = f"실패: {len(results)-pass_count}개"
ws[f'A{summary_row+4}'] = f"성공률: {pass_count/len(results)*100:.1f}%"

# 저장
output_file = f'/Users/mz02-horang/cdrive/test_results_simple_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
wb.save(output_file)

print(f"\n{'='*60}")
print(f"🎉 테스트 완료!")
print(f"📊 결과: {pass_count}/{len(results)} PASS ({pass_count/len(results)*100:.1f}%)")
print(f"📁 파일: {output_file}")
print(f"{'='*60}\n")
