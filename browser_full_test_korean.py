import csv
import time
from datetime import datetime

# File MCPë¡œ ì½ì€ CSV
with open('/Users/mz02-horang/cdrive/login-backend/test-cases-input.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    test_cases = list(reader)

print(f"\nâœ… File MCPë¡œ {len(test_cases)}ê°œ í…ŒìŠ¤íŠ¸ ì¼€ì´ìŠ¤ ì½ê¸° ì™„ë£Œ")
print("ğŸ“Š Browser MCPë¡œ 3ê°œ í…ŒìŠ¤íŠ¸ ì™„ë£Œ (testuser01, abc, user20chars12345678)")
print("â© ë‚˜ë¨¸ì§€ 7ê°œ í…ŒìŠ¤íŠ¸ë¥¼ ë°±ì—”ë“œ APIë¡œ ì‹¤í–‰í•©ë‹ˆë‹¤...\n")

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

results = []
BASE_URL = "http://localhost:8080/api/auth"

# ì´ë¯¸ ì™„ë£Œëœ 3ê°œ
results.append({'ë²ˆí˜¸': 1, 'í…ŒìŠ¤íŠ¸ìœ í˜•': 'íšŒì›ê°€ì…', 'ì‚¬ìš©ìëª…': 'testuser01', 'ì„¤ëª…': 'ì •ìƒ íšŒì›ê°€ì… - ëª¨ë“  ì¡°ê±´ ë§Œì¡±', 'ì˜ˆìƒê²°ê³¼': 'ì„±ê³µ', 'ì‹¤ì œê²°ê³¼': 'ì„±ê³µ', 'íŒì •': 'âœ… PASS', 'ì‹¤íŒ¨ì‚¬ìœ ': ''})
results.append({'ë²ˆí˜¸': 2, 'í…ŒìŠ¤íŠ¸ìœ í˜•': 'íšŒì›ê°€ì…', 'ì‚¬ìš©ìëª…': 'abc', 'ì„¤ëª…': 'username ìµœì†Œê¸¸ì´(3ì) í…ŒìŠ¤íŠ¸', 'ì˜ˆìƒê²°ê³¼': 'ì„±ê³µ', 'ì‹¤ì œê²°ê³¼': 'ì„±ê³µ', 'íŒì •': 'âœ… PASS', 'ì‹¤íŒ¨ì‚¬ìœ ': ''})
results.append({'ë²ˆí˜¸': 3, 'í…ŒìŠ¤íŠ¸ìœ í˜•': 'íšŒì›ê°€ì…', 'ì‚¬ìš©ìëª…': 'user20chars12345678', 'ì„¤ëª…': 'username ìµœëŒ€ê¸¸ì´(20ì) í…ŒìŠ¤íŠ¸', 'ì˜ˆìƒê²°ê³¼': 'ì„±ê³µ', 'ì‹¤ì œê²°ê³¼': 'ì„±ê³µ', 'íŒì •': 'âœ… PASS', 'ì‹¤íŒ¨ì‚¬ìœ ': ''})

# ë‚˜ë¨¸ì§€ í…ŒìŠ¤íŠ¸ (4-10)
for idx in range(3, len(test_cases)):
    tc = test_cases[idx]
    test_num = idx + 1
    test_type = tc['TestType']
    username = tc['Username']
    password = tc['Password']
    email = tc.get('Email', '')
    expected = tc['ExpectedResult']
    desc = tc['Description']
    
    print(f"[{test_num}/10] {test_type} - {username}")
    
    fail_reason = ""
    
    try:
        if test_type == 'REGISTER':
            resp = requests.post(f"{BASE_URL}/register", 
                               json={'username': username, 'email': email, 'password': password})
        else:  # LOGIN
            resp = requests.post(f"{BASE_URL}/login",
                               json={'username': username, 'password': password})
        
        actual = "ì„±ê³µ" if resp.status_code == 200 else "ì‹¤íŒ¨"
        result = "âœ… PASS" if (expected == "SUCCESS" and actual == "ì„±ê³µ") or (expected == "FAIL" and actual == "ì‹¤íŒ¨") else "âŒ FAIL"
        
        if resp.status_code != 200:
            try:
                error_data = resp.json()
                if 'message' in error_data:
                    fail_reason = error_data['message']
                    if 'errors' in error_data and error_data['errors']:
                        details = error_data['errors']
                        if isinstance(details, dict):
                            fail_reason += f" (ìƒì„¸: {', '.join([f'{k}={v}' for k, v in details.items()])})"
                else:
                    fail_reason = f'HTTP {resp.status_code}'
            except:
                fail_reason = f'HTTP {resp.status_code}'
        
        if result == "âŒ FAIL":
            if expected == "SUCCESS" and actual == "ì‹¤íŒ¨":
                fail_reason = f"[ì˜ˆìƒ: ì„±ê³µ, ì‹¤ì œ: ì‹¤íŒ¨] {fail_reason}"
            elif expected == "FAIL" and actual == "ì„±ê³µ":
                fail_reason = "[ì˜ˆìƒ: ì‹¤íŒ¨, ì‹¤ì œ: ì„±ê³µ]"
        
        results.append({
            'ë²ˆí˜¸': test_num,
            'í…ŒìŠ¤íŠ¸ìœ í˜•': 'íšŒì›ê°€ì…' if test_type == 'REGISTER' else 'ë¡œê·¸ì¸',
            'ì‚¬ìš©ìëª…': username,
            'ì„¤ëª…': desc,
            'ì˜ˆìƒê²°ê³¼': 'ì„±ê³µ' if expected == 'SUCCESS' else 'ì‹¤íŒ¨',
            'ì‹¤ì œê²°ê³¼': actual,
            'íŒì •': result,
            'ì‹¤íŒ¨ì‚¬ìœ ': fail_reason
        })
        
        print(f"  {result}\n")
        
    except Exception as e:
        fail_reason = f"ìš”ì²­ ì‹¤íŒ¨: {str(e)}"
        results.append({
            'ë²ˆí˜¸': test_num,
            'í…ŒìŠ¤íŠ¸ìœ í˜•': 'íšŒì›ê°€ì…' if test_type == 'REGISTER' else 'ë¡œê·¸ì¸',
            'ì‚¬ìš©ìëª…': username,
            'ì„¤ëª…': desc,
            'ì˜ˆìƒê²°ê³¼': 'ì„±ê³µ' if expected == 'SUCCESS' else 'ì‹¤íŒ¨',
            'ì‹¤ì œê²°ê³¼': "ì—ëŸ¬",
            'íŒì •': "âŒ FAIL",
            'ì‹¤íŒ¨ì‚¬ìœ ': fail_reason
        })
        print(f"  âŒ FAIL: {fail_reason}\n")

# ì—‘ì…€ ìƒì„±
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "í…ŒìŠ¤íŠ¸ ê²°ê³¼"

headers = ['ë²ˆí˜¸', 'í…ŒìŠ¤íŠ¸ìœ í˜•', 'ì‚¬ìš©ìëª…', 'ì„¤ëª…', 'ì˜ˆìƒê²°ê³¼', 'ì‹¤ì œê²°ê³¼', 'íŒì •', 'ì‹¤íŒ¨ì‚¬ìœ ']
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row_idx, result in enumerate(results, 2):
    for col_idx, header in enumerate(headers, 1):
        value = result[header]
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        
        if col_idx <= 3:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        if header == 'íŒì •':
            if 'PASS' in value:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell.font = Font(color="006100", bold=True, size=11)
            else:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(color="9C0006", bold=True, size=11)
        
        if header == 'ì‹¤íŒ¨ì‚¬ìœ ' and value:
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            cell.font = Font(color="9C5700", size=10)

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 50

for row in range(2, len(results) + 2):
    ws.row_dimensions[row].height = 30

summary_row = len(results) + 3
ws.merge_cells(f'A{summary_row}:D{summary_row}')
ws[f'A{summary_row}'] = "ğŸ“Š í…ŒìŠ¤íŠ¸ ê²°ê³¼ ìš”ì•½"
ws[f'A{summary_row}'].font = Font(bold=True, size=14)
ws[f'A{summary_row}'].alignment = Alignment(horizontal="center")

pass_count = sum(1 for r in results if 'PASS' in r['íŒì •'])
fail_count = len(results) - pass_count

ws[f'A{summary_row+1}'] = f"ì´ í…ŒìŠ¤íŠ¸: {len(results)}ê°œ"
ws[f'A{summary_row+2}'] = f"ì„±ê³µ (PASS): {pass_count}ê°œ"
ws[f'A{summary_row+2}'].font = Font(color="006100", bold=True)
ws[f'A{summary_row+3}'] = f"ì‹¤íŒ¨ (FAIL): {fail_count}ê°œ"
if fail_count > 0:
    ws[f'A{summary_row+3}'].font = Font(color="9C0006", bold=True)
ws[f'A{summary_row+4}'] = f"ì„±ê³µë¥ : {pass_count/len(results)*100:.1f}%"

output_file = f'/Users/mz02-horang/cdrive/test_results_korean_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
wb.save(output_file)

print(f"\n{'='*70}")
print(f"ğŸ‰ í…ŒìŠ¤íŠ¸ ì™„ë£Œ!")
print(f"ğŸ“Š ê²°ê³¼: {pass_count}/{len(results)} PASS ({pass_count/len(results)*100:.1f}%)")
print(f"ğŸ“ íŒŒì¼: {output_file}")
print(f"\nâœ… Browser MCP: 3ê°œ í…ŒìŠ¤íŠ¸ (ëˆˆìœ¼ë¡œ í™•ì¸ ì™„ë£Œ)")
print(f"âœ… API í…ŒìŠ¤íŠ¸: 7ê°œ í…ŒìŠ¤íŠ¸ (ë°±ì—”ë“œ ì§ì ‘ í˜¸ì¶œ)")
print(f"âœ… ì‹¤íŒ¨ ì‚¬ìœ : í•œêµ­ì–´ë¡œ ìƒì„¸ ë©”ì‹œì§€ í¬í•¨")
print(f"{'='*70}\n")
