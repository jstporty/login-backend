import csv
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# File MCP로 읽은 CSV 파싱
with open('/Users/mz02-horang/cdrive/login-backend/test-cases-input.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    test_cases = list(reader)

print(f"\n✅ File MCP로 {len(test_cases)}개 테스트 케이스 읽기 완료\n")

# 테스트 결과
results = []
BASE_URL = "http://localhost:8080/api/auth"

for idx, tc in enumerate(test_cases, 1):
    test_type = tc['TestType']
    username = tc['Username']
    password = tc['Password']
    email = tc.get('Email', '')
    expected = tc['ExpectedResult']
    desc = tc['Description']
    
    print(f"[{idx}/10] {test_type} - {username}")
    
    fail_reason = ""
    
    try:
        if test_type == 'REGISTER':
            resp = requests.post(f"{BASE_URL}/register", 
                               json={'username': username, 'email': email, 'password': password})
        else:  # LOGIN
            resp = requests.post(f"{BASE_URL}/login",
                               json={'username': username, 'password': password})
        
        actual = "성공" if resp.status_code == 200 else "실패"
        result = "✅ PASS" if (expected == "SUCCESS" and actual == "성공") or (expected == "FAIL" and actual == "실패") else "❌ FAIL"
        
        # 실패 사유 추출
        if resp.status_code != 200:
            try:
                error_data = resp.json()
                fail_reason = error_data.get('message', f'HTTP {resp.status_code}')
            except:
                fail_reason = f'HTTP {resp.status_code} - {resp.text[:50]}'
        
        # 예상과 다른 경우
        if result == "❌ FAIL":
            if expected == "SUCCESS" and actual == "실패":
                fail_reason = f"예상: 성공, 실제: 실패 ({fail_reason})"
            elif expected == "FAIL" and actual == "성공":
                fail_reason = "예상: 실패, 실제: 성공"
        
        results.append({
            'No': idx,
            'TestType': '회원가입' if test_type == 'REGISTER' else '로그인',
            'Username': username,
            'Description': desc,
            'Expected': '성공' if expected == 'SUCCESS' else '실패',
            'Actual': actual,
            'Result': result,
            'FailReason': fail_reason
        })
        
        if result == "✅ PASS":
            print(f"  ✅ PASS\n")
        else:
            print(f"  ❌ FAIL: {fail_reason}\n")
        
    except Exception as e:
        fail_reason = f"에러 발생: {str(e)}"
        results.append({
            'No': idx,
            'TestType': '회원가입' if test_type == 'REGISTER' else '로그인',
            'Username': username,
            'Description': desc,
            'Expected': '성공' if expected == 'SUCCESS' else '실패',
            'Actual': "에러",
            'Result': "❌ FAIL",
            'FailReason': fail_reason
        })
        print(f"  ❌ FAIL: {fail_reason}\n")

# 간단한 엑셀 생성
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "테스트 결과"

# 헤더
headers = ['No', 'TestType', 'Username', 'Description', 'Expected', 'Actual', 'Result', 'FailReason']
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 데이터
for row_idx, result in enumerate(results, 2):
    for col_idx, header in enumerate(headers, 1):
        value = result[header]
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 정렬
        if col_idx <= 3:  # No, TestType, Username
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Result 컬럼 색상
        if header == 'Result':
            if 'PASS' in value:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell.font = Font(color="006100", bold=True, size=11)
            else:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(color="9C0006", bold=True, size=11)
        
        # FailReason 컬럼 - 실패한 경우 배경색
        if header == 'FailReason' and value:
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            cell.font = Font(color="9C5700", size=10)

# 컬럼 너비
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 40  # 실패 사유

# 요약 추가
summary_row = len(results) + 3
ws.merge_cells(f'A{summary_row}:D{summary_row}')
ws[f'A{summary_row}'] = "📊 테스트 결과 요약"
ws[f'A{summary_row}'].font = Font(bold=True, size=14)
ws[f'A{summary_row}'].alignment = Alignment(horizontal="center")

pass_count = sum(1 for r in results if 'PASS' in r['Result'])
fail_count = len(results) - pass_count

ws[f'A{summary_row+1}'] = f"총 테스트: {len(results)}개"
ws[f'A{summary_row+2}'] = f"성공 (PASS): {pass_count}개"
ws[f'A{summary_row+2}'].font = Font(color="006100", bold=True)
ws[f'A{summary_row+3}'] = f"실패 (FAIL): {fail_count}개"
if fail_count > 0:
    ws[f'A{summary_row+3}'].font = Font(color="9C0006", bold=True)
ws[f'A{summary_row+4}'] = f"성공률: {pass_count/len(results)*100:.1f}%"

# 저장
output_file = f'/Users/mz02-horang/cdrive/test_results_final_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
wb.save(output_file)

print(f"\n{'='*70}")
print(f"🎉 테스트 완료!")
print(f"📊 결과: {pass_count}/{len(results)} PASS ({pass_count/len(results)*100:.1f}%)")
print(f"📁 파일: {output_file}")
print(f"{'='*70}\n")
