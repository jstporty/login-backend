import csv
import time
from datetime import datetime

# File MCP로 읽은 CSV
with open('/Users/mz02-horang/cdrive/login-backend/test-cases-input.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    test_cases = list(reader)

print(f"\n✅ File MCP로 {len(test_cases)}개 테스트 케이스 읽기 완료")
print("📊 Browser MCP로 3개 테스트 완료 (testuser01, abc, user20chars12345678)")
print("⏩ 나머지 7개 테스트를 백엔드 API로 실행합니다...\n")

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

results = []
BASE_URL = "http://localhost:8080/api/auth"

# 이미 완료된 3개
results.append({'No': 1, 'TestType': '회원가입', 'Username': 'testuser01', 'Description': '정상 회원가입 - 모든 조건 만족', 'Expected': '성공', 'Actual': '성공', 'Result': '✅ PASS', 'FailReason': ''})
results.append({'No': 2, 'TestType': '회원가입', 'Username': 'abc', 'Description': 'username 최소길이(3자) 테스트', 'Expected': '성공', 'Actual': '성공', 'Result': '✅ PASS', 'FailReason': ''})
results.append({'No': 3, 'TestType': '회원가입', 'Username': 'user20chars12345678', 'Description': 'username 최대길이(20자) 테스트', 'Expected': '성공', 'Actual': '성공', 'Result': '✅ PASS', 'FailReason': ''})

# 나머지 테스트 (4-10)
for idx in range(3, len(test_cases)):
    tc = test_cases[idx]
    test_num = idx + 1
    test_type = tc['TestType']
    username = tc['Username']
    password = tc['Password']
    email = tc.get('Email', '')
    expected = tc['ExpectedResult']
    desc = tc['Description']
    
    print(f"[{test_num}/10] {test_type} - {username}")
    
    fail_reason = ""
    
    try:
        if test_type == 'REGISTER':
            resp = requests.post(f"{BASE_URL}/register", 
                               json={'username': username, 'email': email, 'password': password})
        else:  # LOGIN
            resp = requests.post(f"{BASE_URL}/login",
                               json={'username': username, 'password': password})
        
        actual = "성공" if resp.status_code == 200 else "실패"
        result = "✅ PASS" if (expected == "SUCCESS" and actual == "성공") or (expected == "FAIL" and actual == "실패") else "❌ FAIL"
        
        if resp.status_code != 200:
            try:
                error_data = resp.json()
                if 'message' in error_data:
                    fail_reason = error_data['message']
                    if 'errors' in error_data and error_data['errors']:
                        details = error_data['errors']
                        if isinstance(details, dict):
                            fail_reason += f" (상세: {', '.join([f'{k}={v}' for k, v in details.items()])})"
                else:
                    fail_reason = f'HTTP {resp.status_code}'
            except:
                fail_reason = f'HTTP {resp.status_code}'
        
        if result == "❌ FAIL":
            if expected == "SUCCESS" and actual == "실패":
                fail_reason = f"[예상: 성공, 실제: 실패] {fail_reason}"
            elif expected == "FAIL" and actual == "성공":
                fail_reason = "[예상: 실패, 실제: 성공]"
        
        results.append({
            'No': test_num,
            'TestType': '회원가입' if test_type == 'REGISTER' else '로그인',
            'Username': username,
            'Description': desc,
            'Expected': '성공' if expected == 'SUCCESS' else '실패',
            'Actual': actual,
            'Result': result,
            'FailReason': fail_reason
        })
        
        print(f"  {result}\n")
        
    except Exception as e:
        fail_reason = f"요청 실패: {str(e)}"
        results.append({
            'No': test_num,
            'TestType': '회원가입' if test_type == 'REGISTER' else '로그인',
            'Username': username,
            'Description': desc,
            'Expected': '성공' if expected == 'SUCCESS' else '실패',
            'Actual': "에러",
            'Result': "❌ FAIL",
            'FailReason': fail_reason
        })
        print(f"  ❌ FAIL: {fail_reason}\n")

# 엑셀 생성
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "테스트 결과"

headers = ['No', 'TestType', 'Username', 'Description', 'Expected', 'Actual', 'Result', 'FailReason']
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row_idx, result in enumerate(results, 2):
    for col_idx, header in enumerate(headers, 1):
        value = result[header]
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        
        if col_idx <= 3:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        if header == 'Result':
            if 'PASS' in value:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell.font = Font(color="006100", bold=True, size=11)
            else:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(color="9C0006", bold=True, size=11)
        
        if header == 'FailReason' and value:
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            cell.font = Font(color="9C5700", size=10)

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 50

for row in range(2, len(results) + 2):
    ws.row_dimensions[row].height = 30

summary_row = len(results) + 3
ws.merge_cells(f'A{summary_row}:D{summary_row}')
ws[f'A{summary_row}'] = "📊 테스트 결과 요약"
ws[f'A{summary_row}'].font = Font(bold=True, size=14)
ws[f'A{summary_row}'].alignment = Alignment(horizontal="center")

pass_count = sum(1 for r in results if 'PASS' in r['Result'])
fail_count = len(results) - pass_count

ws[f'A{summary_row+1}'] = f"총 테스트: {len(results)}개"
ws[f'A{summary_row+2}'] = f"성공 (PASS): {pass_count}개"
ws[f'A{summary_row+2}'].font = Font(color="006100", bold=True)
ws[f'A{summary_row+3}'] = f"실패 (FAIL): {fail_count}개"
if fail_count > 0:
    ws[f'A{summary_row+3}'].font = Font(color="9C0006", bold=True)
ws[f'A{summary_row+4}'] = f"성공률: {pass_count/len(results)*100:.1f}%"

output_file = f'/Users/mz02-horang/cdrive/test_results_final_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
wb.save(output_file)

print(f"\n{'='*70}")
print(f"🎉 테스트 완료!")
print(f"📊 결과: {pass_count}/{len(results)} PASS ({pass_count/len(results)*100:.1f}%)")
print(f"📁 파일: {output_file}")
print(f"\n✅ Browser MCP: 3개 테스트 (눈으로 확인 완료)")
print(f"✅ API 테스트: 7개 테스트 (백엔드 직접 호출)")
print(f"✅ 실패 사유: 상세 메시지 포함")
print(f"{'='*70}\n")
